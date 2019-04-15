import openpyxl as xl
from openpyxl.chart import BarChart, BarChart3D, BubbleChart, Reference

def process_file(filename):
    read = xl.load_workbook(filename)
    sheet = read['Sheet1']

    for row in range(2, sheet.max_row + 1):
        cell = sheet['c'+str(row)]
        corrected = cell.value * 0.9
        correct_price = sheet['d'+str(row)]
        correct_price.value = corrected

    values = Reference(sheet, min_row=2, max_row=sheet.max_row, min_col= 4, max_col=4)
    chart = BarChart3D()
    chart.add_data(values)
    sheet.add_chart(chart, 'f2')


    read.save(filename)

read = 'transactions2.xlsx'

process_file(read)